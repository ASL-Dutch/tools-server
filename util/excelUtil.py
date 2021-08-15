# -*- coding:utf-8 -*-
import openpyxl
import pandas as pd


# 去读表格
def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames
    print(sheets)
    ws = wb[sheets[0]]

    row = ws.max_row  # 获取最大的 行数
    col = ws.max_column  # 获取最大列数
    line = []
    all_line = []
    num = 0

    for i in range(1, row + 1):
        for x in range(1, col + 1):
            cell = ws.cell(i, x).value
            line.append(cell)
            num += 1
            if num % col == 0:
                all_line.append(line)
                line = []

    return all_line


def read_excel_by_byte(f_byte: bytes):
    wb = openpyxl.load_workbook(f_byte)
    sheets = wb.sheetnames
    print(sheets)
    ws = wb[sheets[0]]

    row = ws.max_row  # 获取最大的 行数
    col = ws.max_column  # 获取最大列数
    line = []
    all_line = []
    num = 0

    for i in range(1, row + 1):
        for x in range(1, col + 1):
            cell = ws.cell(i, x).value
            line.append(cell)
            num += 1
            if num % col == 0:
                all_line.append(line)
                line = []

    return all_line


def read_excel_pandas_by_byte(f_byte: bytes):
    """
    从二进制流中读取excel 内容
    :param f_byte:
    :return:
    """
    df = pd.read_excel(f_byte)
    rows = df.shape[0]
    cols = df.columns.size

    all_line = []
    num = 0
    line = []
    for iRow in range(rows):
        for iCol in range(cols):
            num += 1

            line.append(df.iloc[iRow, iCol])
            print(df.iloc[iRow, iCol])
            if num % cols == 0:
                all_line.append(line)
                line = []

    return all_line
